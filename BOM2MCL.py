# -*- coding:utf-8 -*-

'''
下一版：
将产品名称、工作令报告编号填入表格；
已写入表格的格式与原模板统一
写入文件前验证是否已有重名文件，如有则需重命名后再写入
'''

#定义路径
#BOMPath = r'/home/enzio/PythonPrograms/BOM2MTR'
#MCLSamplePath = r'/home/enzio/PythonPrograms/BOM2MTR/MCL.xlsx'
BOMPath = r'D:\Python\BOM2MCL'
MCLPath = r'D:\Python\BOM2MCL'
MCLSamplePath = r'D:\Python\BOM2MCL\MCL.xlsx'

import os
import re
from openpyxl import load_workbook
import shutil

def loadBOM(jobNo):
    if(not os.access(BOMPath, os.R_OK)):
        print('program try to access', BOMPath,'\nBut access is denied')
        input('Please check the path of BOM in source code and the authorization configuration of BOM floder\nFatal error detected. Program is ending...\nPress ENTER to exit')
        os._exit()
    BOMList = []
    for root, dirs, files in os.walk(BOMPath):
        for f in files:
            if (f.find(jobNo) != -1):
                BOMList.append(os.path.join(root, f))
                print(BOMList)
            else:
                #print('BOM not found in BOMPath, please check the job No. and BOM')
                pass
    return BOMList

def loadMCL(BOMP):
    #check the accessibility of MCL Sample
    #Error in this section, to be solved in next version
    if(not os.access(MCLPath, os.W_OK)):
        print('program try to access', MCLPath,'\nBut access require is denied')
        input('Please check the path of MCL in source code and the authorization configuration of MCL floder\nFatal error detected. Program is ending...\nPress ENTER to exit')
        os._exit()

    #load MCL Sample
    MCLSample = load_workbook(MCLSamplePath)
    #print('MCL Sample:', MCLSample)

    #save MCL Sample as new file
    #use the Job No and in final release
    BOMP_d, BOMP_f = os.path.split(BOMP)
    MCLName = BOMP_f[0:7] + '.xlsx'
    #print(MCLName)
    #Check if the file already exsist before save!!!
    MCLSample.save(os.path.join(MCLPath, MCLName))
    MCL_Path = os.path.join(MCLPath, MCLName)
    #print("MCL Path", MCL_Path, 'MCLPath', MCLPath)
    MCL = load_workbook(MCL_Path)
    return MCL, MCL_Path

def generateMCL(BOMPList):
    for BOMP in BOMPList:
        try:
            wb = load_workbook(BOMP)
        except:
            continue
        MCL_wb, MCL_Path = loadMCL(BOMP)
        MCL = MCL_wb.active
        print('MCL is:', MCL)
        #判断worksheet名是 石墨材料明细表 还是 金属材料明细表 然后确定ws变量指向
        snl = wb.get_sheet_names()
        ws = None
        for sheetname in snl:
            if sheetname == '材料明细表':
                ws = wb.get_sheet_by_name('材料明细表')
            elif sheetname == '石墨材料明细表':
                ws = wb.get_sheet_by_name('石墨材料明细表')
            else:
                #input('Excel文件中未找到材料明细表，请确认文件是否是BOM')
                continue
        if(ws == None):

            continue
        for rowNo in range(11, ws.max_row):
            PartName = ''
            MatSpec = ''
            MatGrade = ''
            PartNo = ''
            if(ws.cell(row = rowNo, column = 2).value == ('★' or '△')):
                print('star or triangle found', rowNo, ws.cell(row = rowNo, column = 2).value)
                try:
                    PartName = str(ws.cell(row = rowNo, column = 4).value)
                    if(PartName == 'None'):
                        PartName = ''
                    print(PartName)
                    MatSpec = str(ws.cell(row = rowNo, column = 5).value)
                    if (MatSpec == 'None'):
                        MatSpec = ''
                    MatGrade = str(ws.cell(row = rowNo, column = 6).value)
                    if(MatGrade == 'None'):
                        MatGrade = ''
                    Size = str(ws.cell(row = rowNo, column = 7).value)
                    if(Size == 'None'):
                        Size = ''
                    PartNo = str(ws.cell(row = rowNo, column = 1).value)
                    if(PartNo == 'None'):
                        PartNo = ''
                    #新增项目放这里
                except:
                    print('error detected when assgning the detail information')
            MCL.cell(row = rowNo-6, column=3).value = PartName
            #'\x0a' 代表换行(new line)，分割材料牌号和等级
            #Add code to handle different situation that if one of MatGrade or MatSpec is missing, there is no need to add '\x0d'
            MCL.cell(row = rowNo-6, column=7).value = (MatGrade+'\x0a'+MatSpec)
            MCL.cell(row = rowNo-6, column=2).value = PartNo
            #MCL save path to be updated with desktop or MTR path
            MCL_wb.save(MCL_Path)
        if(ws == None):
            #Error handling: if the document is not BOM
            #异常处理UnboundLocalError
            input('与输入工作令相关的文件中未找到材料明细表，按ENTER重新查找')
            main()
    return

def main():
    try:
        jobNo = input('This program is to generate MCL using BOM\nPlease input JOB NO:')
        if (jobNo == ('exit' or 'esc' or 'quit' or 'exit()')):
            os._exit()
        else:
            pass
        BOMPList = loadBOM(jobNo)
        generateMCL(BOMPList)
    except Exception as E:
        print('检测到错误：', E)
    finally:
        input('按 ENTER 键退出程序')
        os._exit(0)
main()
